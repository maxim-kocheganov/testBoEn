from ctypes import sizeof
from datetime import datetime
from datetime import date
from django.http import HttpResponse
from django.shortcuts import render
import table.models as m
from django.db.models import Max
from django.shortcuts import redirect
from openpyxl import Workbook,load_workbook
import pytz
import os
from django.conf import settings
from django.shortcuts import redirect

def index(request):
    return render(request, "index.html")

def bd_search(request):
    def fill(i):
        if (i.change >= after and i.change <= before):
            db = {'name':i.name,'create':i.create,\
                'change':i.change,'link':'show/' + str(i.id),\
                'link_download' : 'download/' + str(i.id),\
                'link_id' : str(i.id)}
        return db
    if request.method =="GET":
        name = request.GET.get("db_name", "")
        before = request.GET.get("db_before", "")
        after = request.GET.get("db_after", "")
        change_or_edit = request.GET.get("db_change_or_edit", "change")
        db_order_by = request.GET.get("db_order_by","name")
        if before != "":
            before = datetime.strptime(before, '%d/%m/%Y')
        else:
            before = datetime.now()
        if after != "":
            after =  datetime.strptime(after, '%d/%m/%Y')
        else:
            after = datetime.min
        before = pytz.utc.localize(before)
        after = pytz.utc.localize(after)
        db_order_by = request.GET.get("db_order_by", "name")
        db_order_type = request.GET.get("db_order_type", "ascending")
        dbs1 = []
        dbs_raw = m.DB.objects.all()

        # Order 
        if db_order_by == "changed" and db_order_type == "ascending":
            dbs_raw = dbs_raw.order_by("change")
        elif db_order_by == "changed" and db_order_type == "descending":
            dbs_raw = dbs_raw.order_by("-change")
        if db_order_by == "created" and db_order_type == "ascending":
            dbs_raw = dbs_raw.order_by("create")
        elif db_order_by == "created" and db_order_type == "descending":
            dbs_raw = dbs_raw.order_by("-create")
        if db_order_by == "name" and db_order_type == "ascending":
            dbs_raw = dbs_raw.order_by("name")
        elif db_order_by == "name" and db_order_type == "descending":
            dbs_raw = dbs_raw.order_by("-name")
        
        # Find by name and time range
        if name != "":            
            for i in dbs_raw:
                if (i.name.find(name) != -1):
                    if (change_or_edit == "changed"):
                        if (i.change >= after and i.change <= before):
                            dbs1.append(fill(i))
                    elif (change_or_edit == "created"):
                        if (i.create >= after and i.create <= before):
                           dbs1.append(fill(i))
                    else:
                        dbs1.append(fill(i))
        else:
            for i in dbs_raw:            
                if (change_or_edit == "changed"):
                    if (i.change >= after and i.change <= before):
                        dbs1.append(fill(i))
                elif (change_or_edit == "created"):
                    if (i.create >= after and i.create <= before):
                        dbs1.append(fill(i))
                else:
                    dbs1.append(fill(i))
        
        param = {'db':dbs1}
        return render(request,"search_db.html",param)

def delete(request,id):
    try:
        db = m.DB.objects.filter(id=id)[0]
        db.delete()
    except:
        pass
    return

def db_show(request,id):
    try:
        allCells = m.Cell.objects.filter(db = int(id))
    except:
        return render(request,"404.html")
    #if not allCells:
    #    return render(request,"404.html")
    maxColl = allCells.aggregate(Max('column'))["column__max"]
    maxRow = allCells.aggregate(Max('row'))["row__max"]
    if maxColl is None:
        maxColl = 10
    if maxRow is None:
        maxRow = 10
    res = []
    for i in range(1,maxRow + 1):
        row = []
        for j in range(1,maxColl + 1):
            try:          
                cell = allCells.filter(row = i,column = j)[0]   
                var = cell.Read()      
                if var is not None:
                    row.append({'val':str(var),'column':j, 'row':i})
                else:
                    row.append({'val':"",'column':j, 'row':i})
            except:
                row.append({'val':"",'column':j, 'row':i})       
        res.append(row)

    # Actions for a list
    db = m.DB.objects.filter(id=int(id))
    if not db:
        return render(request,"404.html")
    else:
        db = db[0]
    act = {'name':db.name,'create':db.create,\
                'change':db.change,\
                'link_download' : 'download/' + str(db.id),\
                'link_id' : str(db.id)}
    param = {"res" : res,"act":act,"db_id":db.id}
    return render(request,"show.html", param)

# Process existing or blank db
def process(file):
    db = m.DB()
    if file is not None:
        db.name = file.name 
    else:
        db.name = "Blank.xlsx"
    db.create = date.today()  
    db.change = db.create    
    db.save() 
    if file is not None:
        wb = load_workbook(file)
    else:
        return db.id
    ws = wb.active
    r = 1
    c = 1
    for row in ws.iter_rows():        
        for cell in row:            
            cell = m.Cell()
            cell.row = r
            cell.column = c
            cell.Set(ws._get_cell(r,c).value)
            cell.db = db
            cell.save()
            c += 1
        c = 1
        r += 1
    return db.id

def changeCell(request):
    if request.method =="POST":
        row = request.POST["row"]
        column = request.POST["column"]
        var = request.POST["var"]
        db_id = request.POST["db_id"]
        row = int(row)
        column = int(column)
        db_id = int(db_id)
        try:
            cell = m.Cell.objects.filter(row=row, column=column, db=db_id)[0]
            cell.Set(var)
            cell.save()
        except:
            cell = m.Cell()
            cell.row = row
            cell.column = column
            cell.Set(var)
            db = m.DB.objects.filter(id=db_id)[0]
            cell.db = db
            cell.save()
    return HttpResponse(status=200)
    
def changeName(request):
    if request.method =='POST':
        name = request.POST["name"]
        db_id = request.POST["db_id"]
        db_id = int(db_id)
        db = m.DB.objects.filter(id=db_id)[0]
        db.name = name
        db.save()
        return  redirect("/show/" + str(db_id))

def upload(request):
    if request.method =='POST':
        try:
            file = request.FILES['excel_file']
        except:
            file = None
        db_id = process(file)
        if request.POST["to_edit"]:
            return redirect('/show/' + str(db_id)) # go for an show and edit page
        else:
            return redirect('/')


def download(request, id):
    db = m.DB.objects.filter(id=id)[0]
    wb = Workbook()
    ws = wb.active

    #### create xls
    allCells = m.Cell.objects.filter(db = int(id))
    maxColl = allCells.aggregate(Max('column'))["column__max"]
    maxRow = allCells.aggregate(Max('row'))["row__max"]
    if maxColl is None:
        maxColl = 0
    if maxRow is None:
        maxRow = 0
    for row in range(1,maxRow + 1):
        for coll in range(1,maxColl + 1):
            try:          
                cell = allCells.filter(row = row,column = coll)[0]   
                var = cell.Read()      
                if var is not None:
                    ws.cell(column=coll, row=row, value=var)
                else:
                    pass
            except:
                pass

    #### return xls via http
    name = db.name
    dest_filename = os.path.join(settings.BASE_DIR, 'media',  name)    
    wb.save(filename = dest_filename)
    file = open(dest_filename, 'rb')
    response = HttpResponse(file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename={0}'.format(name)
    return response

# Прервись чтением исходников, уважаемый, лови анекдот:
# Байден выступает перед журналистами:
# — Кто сказал, что я читаю по бумажке? Ха, черточка, ха, черточка, ха!